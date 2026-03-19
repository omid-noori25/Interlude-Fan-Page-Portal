#!/usr/bin/env python3
"""
Interlude Studios Fan Page Portal — Multi-Project Scraper & Site Generator
===========================================================================
Reads project configs from /projects/*.json, scrapes TikTok via RapidAPI,
saves data to /data/, generates per-campaign dashboards in /campaigns/,
and regenerates the portal index.html.

Usage:
  python scraper.py                        # Scrape all active projects
  python scraper.py --project my-project  # Scrape one specific project
  python scraper.py --generate-only       # Rebuild HTML without re-scraping
  python scraper.py --export-xlsx         # Also export Excel files

Setup:
  pip install requests openpyxl
  Set RAPIDAPI_KEY env var, or edit CONFIG below.
"""

import json, time, os, sys, argparse
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests -q")
    import requests

# ─── CONFIG ───────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).parent
PROJECTS_DIR  = ROOT / "projects"
DATA_DIR      = ROOT / "data"
CAMPAIGNS_DIR = ROOT / "campaigns"

CONFIG = {
    "RAPIDAPI_KEY":  os.environ.get("RAPIDAPI_KEY", "96097052a4msh3d8f9e70ed3d27cp17c157jsn71911eb5c3a9"),
    "RAPIDAPI_HOST": "tiktok-scraper7.p.rapidapi.com",
    "USER_POSTS_EP": "https://tiktok-scraper7.p.rapidapi.com/user/posts",
    "REQUEST_DELAY": 2,
}

# ─── BRAND ASSETS ─────────────────────────────────────────────────────────────
LOGO_SVG = (
    '<svg viewBox="0 0 200 220" fill="none" xmlns="http://www.w3.org/2000/svg">'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="104,101 96,114 149,143 157,130"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="96,114 149,143 153,137 100,108"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="96,114 104,101 51,72 43,85"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="104,101 51,72 55,66 108,95"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="96,101 104,114 157,85 149,72"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="104,114 157,85 161,79 108,108"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="104,114 96,101 43,130 51,143"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="96,101 43,130 47,136 100,107"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="92,108 108,108 108,52 92,52"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="108,108 108,52 114,48 114,104"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="108,108 92,108 92,164 108,164"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="92,108 92,164 98,160 98,104"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="100,30 114,38 100,46 86,38"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="100,46 114,54 114,38 100,30"/>'
    '<polygon fill="#000" stroke="white" stroke-width="2.8" stroke-linejoin="round" points="100,46 100,30 86,38 86,54"/>'
    '</svg>'
)

FONT_TAG = (
    '<link rel="preconnect" href="https://fonts.googleapis.com">'
    '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
    '<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;900&display=swap" rel="stylesheet">'
)

def _status_css_class(status):
    return {
        "active":    "status-active",
        "pending":   "status-pending",
        "complete":  "status-complete",
        "cancelled": "status-cancelled",
    }.get(status, "status-pending")

# ─── SCRAPING ─────────────────────────────────────────────────────────────────
def api_headers():
    return {"x-rapidapi-key": CONFIG["RAPIDAPI_KEY"], "x-rapidapi-host": CONFIG["RAPIDAPI_HOST"]}

def scrape_account(username, max_pages=100):
    print(f"    Scraping @{username}...")
    videos, cursor, page = [], "0", 0
    while page < max_pages:
        try:
            r = requests.get(CONFIG["USER_POSTS_EP"], headers=api_headers(),
                params={"unique_id": username, "cursor": cursor, "count": 35}, timeout=30)
            if r.status_code == 429:
                raise Exception("Rate limited")
            r.raise_for_status()
            d = r.json()
            if d.get("code", 0) != 0:
                raise Exception(d.get("msg", "API error"))
            result = d.get("data", {})
            batch  = result.get("videos", [])
            if not batch:
                break
            for v in batch:
                vid_id   = v.get("id") or v.get("video_id")
                author   = v.get("author", {}).get("unique_id", username)
                url      = f"https://www.tiktok.com/@{author}/video/{vid_id}"
                date_str = ""
                ct = v.get("create_time")
                if ct:
                    try:
                        date_str = datetime.fromtimestamp(int(ct)).strftime("%Y-%m-%d")
                    except:
                        pass
                desc = (v.get("title") or "")[:500]
                videos.append({
                    "url": url, "date": date_str,
                    "views": v.get("play_count", 0),    "likes": v.get("digg_count", 0),
                    "comments": v.get("comment_count", 0), "shares": v.get("share_count", 0),
                    "saves": v.get("collect_count", 0),    "downloads": v.get("download_count", 0),
                    "desc": desc, "account": username,
                })
            cursor   = result.get("cursor", "0")
            has_more = result.get("hasMore", False)
            page += 1
            if not has_more:
                break
            time.sleep(CONFIG["REQUEST_DELAY"])
        except Exception as e:
            print(f"      Error on page {page+1}: {e}")
            break
    print(f"      → {len(videos)} videos")
    return videos

def scrape_project(project):
    all_videos = []
    for acct in project["accounts"]:
        videos = scrape_account(acct.lstrip("@"))
        for v in videos:
            v["username"] = "@" + acct.lstrip("@")
        all_videos.extend(videos)
        if acct != project["accounts"][-1]:
            time.sleep(CONFIG["REQUEST_DELAY"])
    return all_videos

# ─── DATA I/O ─────────────────────────────────────────────────────────────────
def save_project_data(project_id, videos):
    DATA_DIR.mkdir(exist_ok=True)
    out = {"scraped_at": datetime.now().isoformat(), "videos": videos}
    (DATA_DIR / f"{project_id}.json").write_text(json.dumps(out, indent=2))

def load_project_data(project_id):
    path = DATA_DIR / f"{project_id}.json"
    if not path.exists():
        return None
    return json.loads(path.read_text())

def load_all_projects():
    projects = []
    for f in sorted(PROJECTS_DIR.glob("*.json")):
        try:
            p = json.loads(f.read_text())
            projects.append(p)
        except Exception as e:
            print(f"  Warning: could not load {f.name}: {e}")
    return projects

# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def export_xlsx(project, videos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system(f"{sys.executable} -m pip install openpyxl -q")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    accounts = list(dict.fromkeys(v["account"] for v in videos))
    wb.remove(wb.active)
    headers = ["Video URL","Date Posted","Views","Likes","Comments","Shares","Saves","Downloads","Description"]
    hfill = PatternFill("solid", fgColor="000000")
    hfont = Font(bold=True, color="FFFFFF")
    for acct in accounts:
        ws = wb.create_sheet(title=f"@{acct[:28]}")
        ws["A1"] = "TikTok Username →"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = acct
        ws["B1"].fill = PatternFill("solid", fgColor="f4f3ef")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal="center")
        acct_vids = sorted([v for v in videos if v["account"] == acct], key=lambda x: x["date"], reverse=True)
        for i, v in enumerate(acct_vids, 3):
            ws.cell(row=i, column=1, value=v["url"])
            ws.cell(row=i, column=2, value=v["date"])
            for j, k in enumerate(["views","likes","comments","shares","saves","downloads"], 3):
                c = ws.cell(row=i, column=j, value=v.get(k, 0))
                c.number_format = "#,##0"
            ws.cell(row=i, column=9, value=v.get("desc", ""))
        for col, w in enumerate([50,18,12,12,12,12,12,12,50], 1):
            ws.column_dimensions[chr(64+col)].width = w
        ws.freeze_panes = "A3"
    out_path = DATA_DIR / f"{project['id']}.xlsx"
    wb.save(out_path)
    print(f"    Excel saved: {out_path}")

# ─── HTML HELPERS ─────────────────────────────────────────────────────────────
def _s(s):
    """Escape string for JS single-quoted string."""
    return s.replace("\\","\\\\").replace("'","\\'").replace("\n"," ").replace("\r","")

def _totals(videos, key):
    return sum(v.get(key, 0) for v in videos)

# ─── CAMPAIGN DASHBOARD GENERATOR ────────────────────────────────────────────
def generate_campaign_html(project, data):
    videos     = data.get("videos", [])
    scraped_at = data.get("scraped_at", "")[:10]
    accounts   = list(dict.fromkeys(v["account"] for v in videos))

    # ── Per-account stats ────────────────────────────────────────────────
    acct_stats = []
    for acct in accounts:
        avs  = [v for v in videos if v["account"] == acct]
        tv   = _totals(avs, "views")
        tl   = _totals(avs, "likes")
        n    = len(avs)
        saves = _totals(avs, "saves")
        dates = sorted(v["date"] for v in avs if v["date"])
        best_views = max((v["views"] for v in avs), default=0)
        eng_rates  = [(v["likes"]+v["comments"]+v["shares"]+v["saves"])/v["views"]*100
                      for v in avs if v["views"] > 0]
        avg_eng    = sum(eng_rates)/len(eng_rates) if eng_rates else 0
        date_range = f"{dates[0]} – {dates[-1]}" if len(dates) >= 2 else (dates[0] if dates else "—")
        acct_stats.append({
            "acct": acct, "n": n, "avs": avs,
            "views": tv, "likes": tl,
            "comments": _totals(avs, "comments"),
            "shares":   _totals(avs, "shares"),
            "saves":    saves,
            "downloads":_totals(avs, "downloads"),
            "like_rate":      f"{tl/tv*100:.1f}" if tv else "0",
            "avg_views":      tv // n if n else 0,
            "best_views":     best_views,
            "avg_eng":        f"{avg_eng:.1f}",
            "saves_per_video":f"{saves/n:.1f}" if n else "0",
            "date_range":     date_range,
        })

    grand = {k: sum(a[k] for a in acct_stats)
             for k in ["views","likes","comments","shares","saves","downloads","n"]}

    # ── JS data arrays per account ───────────────────────────────────────
    js_data_lines = []
    for idx, a in enumerate(acct_stats):
        items = []
        for v in sorted(a["avs"], key=lambda x: x["date"], reverse=True):
            items.append(
                "{url:'" + _s(v["url"]) + "',"
                "date:'" + v["date"] + "',"
                "views:" + str(v["views"]) + ","
                "likes:" + str(v["likes"]) + ","
                "comments:" + str(v["comments"]) + ","
                "shares:" + str(v["shares"]) + ","
                "saves:" + str(v["saves"]) + ","
                "downloads:" + str(v["downloads"]) + ","
                "desc:'" + _s(v["desc"]) + "',"
                "account:'" + _s(v["account"]) + "'}"
            )
        js_data_lines.append(f"const acct{idx}=[{','.join(items)}];")

    acct_spread = ",".join(f"...acct{i}" for i in range(len(acct_stats)))
    js_data_lines.append(
        f"const allRaw=[{acct_spread}].sort((a,b)=>b.date.localeCompare(a.date));"
    )
    js_data_lines.append(
        "const allV=allRaw.slice().sort((a,b)=>b.views-a.views).map((v,i)=>({...v,rank:i+1}));"
    )
    js_data_block = "\n".join(js_data_lines)

    # ── Overview HTML ────────────────────────────────────────────────────
    overview_parts = []
    for a in acct_stats:
        avg_comments  = f"{a['comments']/a['n']:.1f}" if a["n"] else "0"
        avg_downloads = f"{a['downloads']/a['n']:.1f}" if a["n"] else "0"
        save_rate     = f"{a['saves']/a['views']*100:.1f}%" if a["views"] else "0%"
        overview_parts.append(
            f'  <div class="stat-section">\n'
            f'    <div class="stat-section-label">@{a["acct"]}</div>\n'
            f'    <div class="stat-grid">\n'
            f'      <div class="stat-card"><div class="stat-label">Total Views</div><div class="stat-value">{a["views"]:,}</div><div class="stat-sub">{a["n"]} videos · avg {a["avg_views"]:,}</div></div>\n'
            f'      <div class="stat-card"><div class="stat-label">Total Likes</div><div class="stat-value">{a["likes"]:,}</div><div class="stat-sub">{a["like_rate"]}% like rate</div></div>\n'
            f'      <div class="stat-card"><div class="stat-label">Comments</div><div class="stat-value">{a["comments"]:,}</div><div class="stat-sub">avg {avg_comments}/video</div></div>\n'
            f'      <div class="stat-card"><div class="stat-label">Shares</div><div class="stat-value">{a["shares"]:,}</div></div>\n'
            f'      <div class="stat-card"><div class="stat-label">Saves</div><div class="stat-value">{a["saves"]:,}</div><div class="stat-sub">{save_rate} save rate</div></div>\n'
            f'      <div class="stat-card"><div class="stat-label">Downloads</div><div class="stat-value">{a["downloads"]:,}</div><div class="stat-sub">avg {avg_downloads}/video</div></div>\n'
            f'    </div>\n'
            f'  </div>'
        )

    compare_cards = []
    for a in acct_stats:
        compare_cards.append(
            f'    <div class="compare-card">\n'
            f'      <div class="compare-card-title">@{a["acct"]}</div>\n'
            f'      <div class="compare-row"><span class="cl">Videos posted</span><span class="cv">{a["n"]}</span></div>\n'
            f'      <div class="compare-row"><span class="cl">Total views</span><span class="cv">{a["views"]:,}</span></div>\n'
            f'      <div class="compare-row"><span class="cl">Best single video</span><span class="cv">{a["best_views"]:,} views</span></div>\n'
            f'      <div class="compare-row"><span class="cl">Avg engagement rate</span><span class="cv">{a["avg_eng"]}%</span></div>\n'
            f'      <div class="compare-row"><span class="cl">Saves per video</span><span class="cv">{a["saves_per_video"]}</span></div>\n'
            f'      <div class="compare-row"><span class="cl">Date range</span><span class="cv">{a["date_range"]}</span></div>\n'
            f'    </div>'
        )

    overview_html = (
        "\n".join(overview_parts) +
        '\n  <div class="stat-section-label" style="margin-top:32px;margin-bottom:16px">Head-to-Head</div>\n'
        '  <div class="compare-grid">\n' +
        "\n".join(compare_cards) +
        '\n  </div>'
    )

    # ── Trends section HTML ──────────────────────────────────────────────
    daily_cards = []
    for idx, a in enumerate(acct_stats):
        daily_cards.append(
            f'      <div class="chart-card">'
            f'<div class="chart-title">@{a["acct"]} — Daily Views</div>'
            f'<div class="chart-sub">Sum of views per day</div>'
            f'<canvas id="dc{idx}" height="200"></canvas></div>'
        )
    trends_daily_html = '    <div class="chart-grid two">\n' + "\n".join(daily_cards) + '\n    </div>'

    # ── Table filter buttons ─────────────────────────────────────────────
    tbl_btns = ['<button class="tfbtn active" onclick="fTbl(\'all\',this)">All</button>']
    for a in acct_stats:
        tbl_btns.append(f'<button class="tfbtn" onclick="fTbl(\'{a["acct"]}\',this)">@{a["acct"]}</button>')
    tbl_filter_html = '<div class="tfilter-group">' + "".join(tbl_btns) + '</div>'

    # ── Posts filter buttons ─────────────────────────────────────────────
    post_btns = [f'<button class="pfbtn active" onclick="fPosts(\'all\',this)">All ({grand["n"]})</button>']
    for a in acct_stats:
        post_btns.append(f'<button class="pfbtn" onclick="fPosts(\'{a["acct"]}\',this)">@{a["acct"]} ({a["n"]})</button>')
    post_filter_html = '<div class="post-filters">' + "".join(post_btns) + '</div>'

    # ── Engagement section HTML ──────────────────────────────────────────
    donut_cards = []
    stacked_cards = []
    for idx, a in enumerate(acct_stats):
        donut_cards.append(
            f'      <div class="chart-card">'
            f'<div class="chart-title">Engagement Mix — @{a["acct"]}</div>'
            f'<div class="chart-sub">Share of total engagement by type</div>'
            f'<div style="max-width:280px;margin:0 auto"><canvas id="dn{idx}" height="240"></canvas></div></div>'
        )
        stacked_cards.append(
            f'    <div class="chart-grid one" style="margin-bottom:1px">'
            f'<div class="chart-card">'
            f'<div class="chart-title">Engagement Breakdown — @{a["acct"]}</div>'
            f'<div class="chart-sub">Stacked: likes · comments · shares · saves · downloads</div>'
            f'<canvas id="sb{idx}" height="240"></canvas></div></div>'
        )
    donut_html   = '    <div class="chart-grid two" style="margin-bottom:1px">\n' + "\n".join(donut_cards) + '\n    </div>'
    stacked_html = "\n".join(stacked_cards)

    # ── Posts summary ────────────────────────────────────────────────────
    posts_summary_html = (
        '    <div class="posts-summary">'
        f'<div class="psum"><div class="psum-val">{grand["n"]}</div><div class="psum-lbl">Total Posts</div></div>'
        f'<div class="psum"><div class="psum-val">{grand["views"]:,}</div><div class="psum-lbl">Total Views</div></div>'
        f'<div class="psum"><div class="psum-val">{grand["likes"]:,}</div><div class="psum-lbl">Total Likes</div></div>'
        f'<div class="psum"><div class="psum-val">{grand["shares"]:,}</div><div class="psum-lbl">Total Shares</div></div>'
        f'<div class="psum"><div class="psum-val">{grand["saves"]:,}</div><div class="psum-lbl">Total Saves</div></div>'
        '</div>'
    )

    # ── Dynamic JS functions ─────────────────────────────────────────────
    # bTrend datasets
    trend_datasets = []
    for idx, a in enumerate(acct_stats):
        vn   = f"acct{idx}"
        name = f"@{a['acct']}" if idx == 0 else a["acct"]
        if idx == 0:
            ds = (
                "{label:'" + name + "',"
                "data:[..." + vn + "].sort((a,b)=>a.date.localeCompare(b.date)).map(v=>v.views),"
                "borderColor:'#000',backgroundColor:'rgba(0,0,0,0.06)',"
                "borderWidth:2,pointRadius:3,pointBackgroundColor:'#000',tension:0.3,fill:true}"
            )
        else:
            ds = (
                "{label:'" + name + "',"
                "data:[..." + vn + "].sort((a,b)=>a.date.localeCompare(b.date)).map(v=>v.views),"
                "borderColor:'#888',backgroundColor:'rgba(136,136,136,0.04)',"
                "borderWidth:1.5,pointRadius:3,pointBackgroundColor:'#888',"
                "tension:0.3,fill:true,borderDash:[4,3]}"
            )
        trend_datasets.append(ds)

    daily_inits = "\n  ".join(
        f"const _nd{i}=gbd(acct{i},'views');"
        f"mkBar('dc{i}',Object.keys(_nd{i}),Object.values(_nd{i}));"
        for i in range(len(acct_stats))
    )
    setM_updates = "\n  ".join(
        f"tco.data.datasets[{i}].data="
        f"[...acct{i}].sort((a,b)=>a.date.localeCompare(b.date)).map(v=>v[m]);"
        for i in range(len(acct_stats))
    )
    donut_inits = "\n  ".join(
        f"new Chart(document.getElementById('dn{i}'),"
        f"dOpts(['Likes','Comments','Shares','Saves','Downloads'],"
        f"[sm(acct{i},'likes'),sm(acct{i},'comments'),sm(acct{i},'shares'),sm(acct{i},'saves'),sm(acct{i},'downloads')]));"
        for i in range(len(acct_stats))
    )
    stacked_inits = "\n  ".join(
        f"mkSB('sb{i}',[...acct{i}].sort((a,b)=>a.date.localeCompare(b.date)));"
        for i in range(len(acct_stats))
    )
    trend_datasets_js = ",".join(trend_datasets)

    # ── Metadata ─────────────────────────────────────────────────────────
    status      = project.get("status", "active")
    status_css  = _status_css_class(status)
    accts_disp  = " · ".join("@" + a.lstrip("@") for a in project["accounts"][:4])
    proj_name   = project["name"]

    # ── CSS (static — defined as plain string, no f-string interpretation) ──
    CAMPAIGN_CSS = """  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --black: #000000; --white: #ffffff; --off: #f4f3ef;
    --gray1: #1a1a1a; --gray2: #333333; --gray3: #888888; --gray4: #cccccc; --gray5: #e8e8e4;
    --font: 'Inter', -apple-system, BlinkMacSystemFont, 'Helvetica Neue', Arial, sans-serif;
  }
  html { background: var(--off); }
  body { font-family: var(--font); color: var(--black); min-height: 100vh; }
  a { color: inherit; text-decoration: none; }
  header { background: var(--black); padding: 0 40px; display: flex; align-items: center; gap: 24px; height: 64px; border-bottom: 1px solid #1a1a1a; position: sticky; top: 0; z-index: 200; }
  .back-btn { font-size: 11px; font-weight: 500; letter-spacing: 0.14em; text-transform: uppercase; color: var(--gray3); padding: 8px 0; border-bottom: 1px solid transparent; transition: color .15s, border-color .15s; white-space: nowrap; flex-shrink: 0; }
  .back-btn:hover { color: var(--white); border-bottom-color: var(--white); }
  .header-divider { width: 1px; height: 24px; background: #2a2a2a; flex-shrink: 0; }
  .header-logo { display: flex; align-items: center; gap: 10px; flex-shrink: 0; }
  .header-logo svg { width: 22px; height: 22px; }
  .header-logo-text { font-size: 11px; font-weight: 700; letter-spacing: 0.18em; text-transform: uppercase; color: var(--white); }
  .header-campaign { flex: 1; display: flex; align-items: center; gap: 12px; overflow: hidden; }
  .header-campaign-name { font-size: 13px; font-weight: 700; letter-spacing: 0.04em; color: var(--white); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .header-status { font-size: 9px; font-weight: 600; letter-spacing: 0.2em; text-transform: uppercase; padding: 3px 8px; border: 1px solid var(--white); color: var(--white); flex-shrink: 0; }
  .share-btn { margin-left: auto; padding: 9px 18px; border: 1px solid #333; color: var(--gray3); font-family: var(--font); font-size: 10px; font-weight: 600; letter-spacing: 0.16em; text-transform: uppercase; cursor: pointer; background: transparent; transition: all .15s; flex-shrink: 0; }
  .share-btn:hover { border-color: var(--white); color: var(--white); }
  .campaign-bar { background: var(--black); border-bottom: 1px solid #111; padding: 22px 40px 20px; }
  .campaign-title { font-size: 38px; font-weight: 900; letter-spacing: -0.02em; color: var(--white); text-transform: uppercase; line-height: 1; }
  .campaign-meta { font-size: 11px; font-weight: 400; letter-spacing: 0.12em; text-transform: uppercase; color: var(--gray3); margin-top: 8px; display: flex; align-items: center; gap: 16px; }
  .meta-sep { color: #333; }
  .tabs { background: var(--black); border-bottom: 1px solid #111; padding: 0 40px; display: flex; align-items: stretch; position: sticky; top: 64px; z-index: 100; }
  .tab { padding: 14px 20px; font-size: 11px; font-weight: 500; letter-spacing: 0.14em; text-transform: uppercase; color: var(--gray3); cursor: pointer; border-bottom: 2px solid transparent; transition: color .15s, border-color .15s; white-space: nowrap; }
  .tab:hover { color: var(--white); }
  .tab.active { color: var(--white); border-bottom-color: var(--white); }
  .content { padding: 0 40px 60px; }
  .section { display: none; }
  .section.active { display: block; }
  .stat-section { margin-top: 32px; margin-bottom: 32px; }
  .stat-section-label { font-size: 10px; font-weight: 600; letter-spacing: 0.22em; text-transform: uppercase; color: var(--gray3); margin-bottom: 12px; padding-bottom: 8px; border-bottom: 1px solid var(--gray5); display: flex; align-items: center; gap: 10px; }
  .stat-section-label::before { content: ''; flex: 1; }
  .stat-section-label::after  { content: ''; flex: 1; }
  .stat-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 1px; background: var(--black); border: 1px solid var(--black); }
  .stat-card { background: var(--white); padding: 20px 18px 18px; }
  .stat-label { font-size: 9px; font-weight: 600; letter-spacing: 0.2em; text-transform: uppercase; color: var(--gray3); margin-bottom: 10px; }
  .stat-value { font-size: 28px; font-weight: 900; letter-spacing: -0.03em; line-height: 1; }
  .stat-sub { font-size: 10px; color: var(--gray3); margin-top: 5px; letter-spacing: 0.06em; }
  .compare-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(300px,1fr)); gap: 1px; background: var(--black); border: 1px solid var(--black); margin-bottom: 32px; }
  .compare-card { background: var(--white); padding: 24px; }
  .compare-card-title { font-size: 12px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; margin-bottom: 18px; border-bottom: 1px solid var(--black); padding-bottom: 12px; }
  .compare-row { display: flex; justify-content: space-between; align-items: baseline; padding: 8px 0; border-bottom: 1px solid var(--gray5); font-size: 13px; }
  .compare-row:last-child { border-bottom: none; }
  .compare-row .cl { color: var(--gray3); font-size: 11px; letter-spacing: 0.06em; }
  .compare-row .cv { font-weight: 700; }
  .chart-grid { display: grid; gap: 1px; background: var(--black); border: 1px solid var(--black); }
  .chart-grid.two { grid-template-columns: 1fr 1fr; }
  .chart-grid.one { grid-template-columns: 1fr; }
  .chart-card { background: var(--white); padding: 24px; }
  .chart-title { font-size: 11px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; margin-bottom: 4px; }
  .chart-sub { font-size: 10px; color: var(--gray3); letter-spacing: 0.06em; margin-bottom: 20px; }
  .metric-row { display: flex; gap: 0; margin-bottom: 20px; border: 1px solid var(--black); width: fit-content; }
  .mpill { padding: 7px 14px; font-size: 10px; font-weight: 600; letter-spacing: 0.14em; text-transform: uppercase; cursor: pointer; color: var(--gray3); background: var(--white); border-right: 1px solid var(--black); transition: all .15s; }
  .mpill:last-child { border-right: none; }
  .mpill:hover { background: var(--off); color: var(--black); }
  .mpill.active { background: var(--black); color: var(--white); }
  .table-wrap { border: 1px solid var(--black); background: var(--black); margin-bottom: 32px; }
  .table-head-bar { background: var(--white); padding: 18px 20px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 12px; border-bottom: 1px solid var(--black); }
  .table-head-bar h3 { font-size: 11px; font-weight: 700; letter-spacing: 0.16em; text-transform: uppercase; }
  .tfilter-group { display: flex; gap: 0; border: 1px solid var(--black); }
  .tfbtn { padding: 7px 14px; font-size: 10px; font-weight: 600; letter-spacing: 0.12em; text-transform: uppercase; cursor: pointer; color: var(--gray3); background: var(--white); border-right: 1px solid var(--black); font-family: var(--font); transition: all .15s; }
  .tfbtn:last-child { border-right: none; }
  .tfbtn:hover { background: var(--off); color: var(--black); }
  .tfbtn.active { background: var(--black); color: var(--white); }
  .table-scroll { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; background: var(--white); }
  th { text-align: left; padding: 10px 14px; font-size: 9px; font-weight: 600; letter-spacing: 0.18em; text-transform: uppercase; color: var(--gray3); background: var(--off); cursor: pointer; user-select: none; border-bottom: 1px solid var(--black); white-space: nowrap; }
  th:hover { color: var(--black); }
  td { padding: 11px 14px; font-size: 12px; border-bottom: 1px solid var(--gray5); }
  tr:hover td { background: var(--off); }
  .td-desc { max-width: 260px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; color: var(--gray3); font-size: 11px; }
  .mb { display: flex; align-items: center; gap: 8px; }
  .bbg { flex: 1; height: 2px; background: var(--gray5); }
  .bf { height: 100%; background: var(--black); }
  .acct-tag { font-size: 9px; font-weight: 600; letter-spacing: 0.14em; text-transform: uppercase; padding: 3px 7px; border: 1px solid var(--black); }
  .posts-summary { display: flex; gap: 0; border: 1px solid var(--black); background: var(--black); margin-bottom: 20px; }
  .psum { background: var(--white); flex: 1; padding: 16px 18px; border-right: 1px solid var(--black); }
  .psum:last-child { border-right: none; }
  .psum-val { font-size: 20px; font-weight: 900; letter-spacing: -0.02em; }
  .psum-lbl { font-size: 9px; font-weight: 600; letter-spacing: 0.18em; text-transform: uppercase; color: var(--gray3); margin-top: 3px; }
  .post-filters { display: flex; gap: 0; border: 1px solid var(--black); width: fit-content; margin-bottom: 20px; }
  .pfbtn { padding: 9px 16px; font-size: 10px; font-weight: 600; letter-spacing: 0.12em; text-transform: uppercase; cursor: pointer; color: var(--gray3); background: var(--white); border-right: 1px solid var(--black); font-family: var(--font); transition: all .15s; }
  .pfbtn:last-child { border-right: none; }
  .pfbtn:hover { background: var(--off); color: var(--black); }
  .pfbtn.active { background: var(--black); color: var(--white); }
  .posts-note { font-size: 10px; letter-spacing: 0.08em; color: var(--gray3); padding: 10px 14px; border: 1px solid var(--gray5); background: var(--white); margin-bottom: 20px; text-transform: uppercase; }
  .pgrid { display: grid; grid-template-columns: repeat(auto-fill, minmax(168px, 1fr)); gap: 1px; background: var(--black); border: 1px solid var(--black); }
  .pcard { background: var(--white); display: flex; flex-direction: column; cursor: pointer; transition: background .12s; text-decoration: none; color: inherit; }
  .pcard:hover { background: var(--off); }
  .pthumb { position: relative; width: 100%; aspect-ratio: 9/16; background: var(--black); overflow: hidden; flex-shrink: 0; }
  .pthumb img { width: 100%; height: 100%; object-fit: cover; display: block; transition: opacity .3s; }
  .pthumb .ph { width: 100%; height: 100%; display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 6px; }
  .pthumb .ph .tl { font-size: 2rem; color: #333; }
  .pthumb .ph .lt { font-size: 9px; letter-spacing: 0.12em; text-transform: uppercase; color: #444; }
  .pthumb .pso { position: absolute; bottom: 0; left: 0; right: 0; background: linear-gradient(transparent, rgba(0,0,0,0.9)); padding: 20px 8px 8px; display: flex; gap: 8px; }
  .ostat { font-size: 9px; font-weight: 600; color: var(--white); letter-spacing: 0.04em; }
  .pplat { position: absolute; top: 8px; left: 8px; font-size: 10px; font-weight: 700; letter-spacing: 0.1em; color: var(--white); background: rgba(0,0,0,0.7); padding: 3px 6px; text-transform: uppercase; }
  .pfoot { padding: 10px 10px 12px; display: flex; flex-direction: column; gap: 4px; flex: 1; border-top: 1px solid var(--black); }
  .puser { font-size: 10px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; }
  .pdesc { font-size: 10px; color: var(--gray3); line-height: 1.4; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }
  .pstats { display: flex; gap: 6px; margin-top: 4px; flex-wrap: wrap; }
  .pchip { font-size: 9px; color: var(--gray3); letter-spacing: 0.06em; }
  .pdate { font-size: 9px; letter-spacing: 0.1em; text-transform: uppercase; color: var(--gray4); margin-top: 2px; }
  @keyframes spin { to { transform: rotate(360deg); } }
  .sp { width: 16px; height: 16px; border: 1.5px solid #333; border-top-color: var(--white); border-radius: 50%; animation: spin .8s linear infinite; }
  @media (max-width: 768px) {
    header, .campaign-bar, .tabs, .content { padding-left: 20px; padding-right: 20px; }
    .campaign-title { font-size: 26px; }
    .chart-grid.two { grid-template-columns: 1fr; }
    .compare-grid { grid-template-columns: 1fr; }
    .posts-summary { flex-wrap: wrap; }
    .pgrid { grid-template-columns: repeat(auto-fill, minmax(130px,1fr)); }
  }"""

    # ── Static JS rendering functions (plain string — no f-string) ──────
    JS_STATIC = """Chart.defaults.font.family = "'Inter', sans-serif";
Chart.defaults.font.size   = 11;
const GRID_COLOR  = '#e8e8e4';
const LABEL_COLOR = '#888';
const scaleOpts = {
  x: { ticks:{color:LABEL_COLOR,maxTicksLimit:10}, grid:{color:GRID_COLOR}, border:{color:'#000'} },
  y: { ticks:{color:LABEL_COLOR,callback:v=>v>=1000?(v/1000).toFixed(0)+'K':v}, grid:{color:GRID_COLOR}, border:{color:'#000'} },
};
const built = {tr:false,tb:false,en:false,po:false};
function sw(n,el) {
  document.querySelectorAll('.section').forEach(s=>s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('tab-'+n).classList.add('active');
  el.classList.add('active');
  if(n==='trends'&&!built.tr){bTrend();built.tr=true;}
  if(n==='videos'&&!built.tb){rTbl();built.tb=true;}
  if(n==='engagement'&&!built.en){bEng();built.en=true;}
  if(n==='posts'&&!built.po){bPosts(allRaw.slice());built.po=true;}
}
const tc={};
function fn(n){return n>=1e6?(n/1e6).toFixed(1)+'M':n>=1e3?(n/1e3).toFixed(1)+'K':n.toString();}
function bPosts(data) {
  const g=document.getElementById('pg'); g.innerHTML='';
  data.forEach(v=>{
    const id='t'+v.url.split('/').pop();
    const card=document.createElement('a');
    card.className='pcard'; card.href=v.url; card.target='_blank'; card.rel='noopener';
    card.innerHTML=`<div class="pthumb" id="${id}"><div class="ph"><div class="tl">*</div><div class="sp"></div><div class="lt">Loading</div></div><div class="pplat">TT</div><div class="pso"><span class="ostat">▶ ${fn(v.views)}</span><span class="ostat">♥ ${fn(v.likes)}</span></div></div><div class="pfoot"><div class="puser">${v.account}</div><div class="pdesc">${v.desc}</div><div class="pstats"><span class="pchip">↩ ${v.shares}</span><span class="pchip">◈ ${v.saves}</span></div><div class="pdate">${v.date}</div></div>`;
    g.appendChild(card);
    lThumb(v.url,id);
  });
}
async function lThumb(url,id){
  if(tc[url]){aThumb(id,tc[url]);return;}
  try{
    const r=await fetch('https://www.tiktok.com/oembed?url='+encodeURIComponent(url));
    if(!r.ok) throw 0;
    const d=await r.json();
    if(d.thumbnail_url){tc[url]=d.thumbnail_url;aThumb(id,d.thumbnail_url);}else sFb(id);
  }catch(e){sFb(id);}
}
function aThumb(id,src){
  const c=document.getElementById(id); if(!c) return;
  const p=c.querySelector('.ph'); if(p) p.remove();
  const img=document.createElement('img'); img.src=src; img.alt='';
  img.style.opacity='0'; img.onload=()=>img.style.opacity='1';
  img.onerror=()=>{sFb(id);img.remove();};
  c.insertBefore(img,c.querySelector('.pso')||c.querySelector('.pplat'));
}
function sFb(id){
  const c=document.getElementById(id); if(!c) return;
  const p=c.querySelector('.ph');
  if(p) p.innerHTML='<div class="tl" style="opacity:.15">*</div><div class="lt">Unavailable</div>';
}
function fPosts(f,btn){
  document.querySelectorAll('.pfbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  bPosts(f==='all'?allRaw.slice():allRaw.filter(v=>v.account===f));
}
let cf='all',sc='views',sa=false;
function fTbl(f,btn){cf=f;document.querySelectorAll('.tfbtn').forEach(b=>b.classList.remove('active'));btn.classList.add('active');rTbl();}
function srt(col){if(sc===col)sa=!sa;else{sc=col;sa=col==='date';};rTbl();}
function rTbl(){
  let data=cf==='all'?allV:allV.filter(v=>v.account===cf);
  data=[...data].sort((a,b)=>{let va=a[sc],vb=b[sc];if(typeof va==='string')return sa?va.localeCompare(vb):vb.localeCompare(va);return sa?va-vb:vb-va;});
  const mx=Math.max(...allV.map(v=>v.views));
  document.getElementById('vtb').innerHTML=data.map((v,i)=>`<tr><td style="font-weight:700;color:#888;font-size:11px">${i+1}</td><td><span class="acct-tag">${v.account}</span></td><td style="color:#888;white-space:nowrap;font-size:11px">${v.date}</td><td><div class="mb"><span style="min-width:60px;font-weight:700;font-size:13px">${v.views.toLocaleString()}</span><div class="bbg"><div class="bf" style="width:${(v.views/mx*100).toFixed(1)}%"></div></div></div></td><td style="font-weight:600">${v.likes.toLocaleString()}</td><td>${v.comments}</td><td>${v.shares}</td><td>${v.saves}</td><td class="td-desc" title="${v.desc}">${v.desc}</td></tr>`).join('');
}
let tco=null;
const mL={views:'Views Over Time',likes:'Likes Over Time',comments:'Comments Over Time',shares:'Shares Over Time',saves:'Saves Over Time'};
function gbd(arr,m){const o={};arr.slice().sort((a,b)=>a.date.localeCompare(b.date)).forEach(v=>{o[v.date]=(o[v.date]||0)+v[m];});return o;}
function mkBar(id,labels,data){
  new Chart(document.getElementById(id),{type:'bar',data:{labels,datasets:[{data,backgroundColor:'#000',borderColor:'#000',borderWidth:0,borderRadius:0}]},options:{responsive:true,plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.y.toLocaleString()}`}}},scales:scaleOpts}});
}
function sm(arr,k){return arr.reduce((s,v)=>s+v[k],0);}"""

    # Assemble final HTML
    html = (
        f'<!DOCTYPE html>\n<html lang="en">\n<head>\n'
        f'<meta charset="UTF-8">\n'
        f'<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        f'<title>{proj_name} — Interlude Studios</title>\n'
        f'{FONT_TAG}\n'
        f'<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>\n'
        f'<style>\n{CAMPAIGN_CSS}\n</style>\n'
        f'</head>\n<body>\n\n'
        # Header
        f'<header>\n'
        f'  <a class="back-btn" href="../index.html">← All Campaigns</a>\n'
        f'  <div class="header-divider"></div>\n'
        f'  <div class="header-logo">{LOGO_SVG}<span class="header-logo-text">Interlude Studios</span></div>\n'
        f'  <div class="header-divider"></div>\n'
        f'  <div class="header-campaign">\n'
        f'    <span class="header-campaign-name">{proj_name}</span>\n'
        f'    <span class="header-status">{status.title()}</span>\n'
        f'  </div>\n'
        f'  <button class="share-btn" onclick="navigator.clipboard&&navigator.clipboard.writeText(location.href)">Share ↗</button>\n'
        f'</header>\n\n'
        # Campaign bar
        f'<div class="campaign-bar">\n'
        f'  <div class="campaign-title">{proj_name} +</div>\n'
        f'  <div class="campaign-meta">\n'
        f'    <span>TikTok</span><span class="meta-sep">·</span>\n'
        f'    <span>{accts_disp}</span><span class="meta-sep">·</span>\n'
        f'    <span>Last updated: {scraped_at}</span>\n'
        f'  </div>\n'
        f'</div>\n\n'
        # Tabs
        f'<nav class="tabs">\n'
        f'  <div class="tab active" onclick="sw(\'overview\',this)">Overview</div>\n'
        f'  <div class="tab" onclick="sw(\'trends\',this)">Trends</div>\n'
        f'  <div class="tab" onclick="sw(\'videos\',this)">Top Videos</div>\n'
        f'  <div class="tab" onclick="sw(\'engagement\',this)">Engagement</div>\n'
        f'  <div class="tab" onclick="sw(\'posts\',this)">Posts</div>\n'
        f'</nav>\n\n'
        # Content
        f'<div class="content">\n\n'
        # Overview
        f'<div id="tab-overview" class="section active">\n'
        f'{overview_html}\n'
        f'</div>\n\n'
        # Trends
        f'<div id="tab-trends" class="section">\n'
        f'  <div style="margin-top:28px">\n'
        f'    <div class="metric-row">\n'
        f'      <div class="mpill active" data-m="views" onclick="setM(\'views\')">Views</div>\n'
        f'      <div class="mpill" data-m="likes" onclick="setM(\'likes\')">Likes</div>\n'
        f'      <div class="mpill" data-m="comments" onclick="setM(\'comments\')">Comments</div>\n'
        f'      <div class="mpill" data-m="shares" onclick="setM(\'shares\')">Shares</div>\n'
        f'      <div class="mpill" data-m="saves" onclick="setM(\'saves\')">Saves</div>\n'
        f'    </div>\n'
        f'    <div class="chart-grid one" style="margin-bottom:1px">\n'
        f'      <div class="chart-card">\n'
        f'        <div class="chart-title" id="tt">Views Over Time</div>\n'
        f'        <div class="chart-sub">Each point = one video · sorted by date posted</div>\n'
        f'        <canvas id="trendChart" height="280"></canvas>\n'
        f'      </div>\n'
        f'    </div>\n'
        f'{trends_daily_html}\n'
        f'  </div>\n'
        f'</div>\n\n'
        # Top Videos
        f'<div id="tab-videos" class="section">\n'
        f'  <div class="table-wrap" style="margin-top:28px">\n'
        f'    <div class="table-head-bar">\n'
        f'      <h3>All Videos — Ranked</h3>\n'
        f'      {tbl_filter_html}\n'
        f'    </div>\n'
        f'    <div class="table-scroll"><table>\n'
        f'      <thead><tr>\n'
        f'        <th onclick="srt(\'rank\')">#</th>\n'
        f'        <th>Account</th>\n'
        f'        <th onclick="srt(\'date\')">Date ↕</th>\n'
        f'        <th onclick="srt(\'views\')">Views ↓</th>\n'
        f'        <th onclick="srt(\'likes\')">Likes ↕</th>\n'
        f'        <th onclick="srt(\'comments\')">Comments ↕</th>\n'
        f'        <th onclick="srt(\'shares\')">Shares ↕</th>\n'
        f'        <th onclick="srt(\'saves\')">Saves ↕</th>\n'
        f'        <th>Description</th>\n'
        f'      </tr></thead>\n'
        f'      <tbody id="vtb"></tbody>\n'
        f'    </table></div>\n'
        f'  </div>\n'
        f'</div>\n\n'
        # Engagement
        f'<div id="tab-engagement" class="section">\n'
        f'  <div style="margin-top:28px">\n'
        f'{donut_html}\n'
        f'{stacked_html}\n'
        f'  </div>\n'
        f'</div>\n\n'
        # Posts
        f'<div id="tab-posts" class="section">\n'
        f'  <div style="margin-top:28px">\n'
        f'{posts_summary_html}\n'
        f'    {post_filter_html}\n'
        f'    <div class="posts-note">Thumbnails load from TikTok when opened with internet · Click any card to open the original post</div>\n'
        f'    <div class="pgrid" id="pg"></div>\n'
        f'  </div>\n'
        f'</div>\n\n'
        f'</div><!-- /content -->\n\n'
        # Scripts
        f'<script>\n'
        f'{js_data_block}\n\n'
        f'{JS_STATIC}\n\n'
        # bTrend (dynamic)
        f'function bTrend(){{\n'
        f'  const labels=[...acct0].sort((a,b)=>a.date.localeCompare(b.date)).map(v=>v.date);\n'
        f'  tco=new Chart(document.getElementById(\'trendChart\'),{{type:\'line\',\n'
        f'    data:{{labels,datasets:[{trend_datasets_js}]}},\n'
        f'    options:{{responsive:true,interaction:{{mode:\'index\',intersect:false}},\n'
        f'      plugins:{{legend:{{labels:{{color:\'#333\',boxWidth:12,font:{{size:11,weight:\'500\'}}}}}},\n'
        f'        tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y.toLocaleString()}}`}}}}}},\n'
        f'      scales:scaleOpts}}}});\n'
        f'  {daily_inits}\n'
        f'}}\n\n'
        # setM (dynamic)
        f'function setM(m){{\n'
        f'  document.querySelectorAll(\'.mpill\').forEach(p=>p.classList.toggle(\'active\',p.dataset.m===m));\n'
        f'  document.getElementById(\'tt\').textContent=mL[m];\n'
        f'  if(!tco)return;\n'
        f'  {setM_updates}\n'
        f'  tco.update();\n'
        f'}}\n\n'
        # bEng (dynamic)
        f'function bEng(){{\n'
        f"  const ENG_COLORS=['#000000','#444444','#777777','#aaaaaa','#cccccc'];\n"
        f'  const dOpts=(l,d)=>({{type:\'doughnut\',data:{{labels:l,datasets:[{{data:d,backgroundColor:ENG_COLORS,borderColor:\'#f4f3ef\',borderWidth:3,hoverOffset:6}}]}},\n'
        f'    options:{{responsive:true,plugins:{{legend:{{position:\'bottom\',labels:{{color:\'#333\',padding:14,boxWidth:10,font:{{size:10,weight:\'600\'}}}}}},\n'
        f'      tooltip:{{callbacks:{{label:ctx=>` ${{ctx.label}}: ${{ctx.parsed.toLocaleString()}} (${{(ctx.parsed/ctx.dataset.data.reduce((a,b)=>a+b)*100).toFixed(1)}}%)`}}}}}}}}}});\n'
        f'  function mkSB(id,data){{\n'
        f'    const labels=data.map((v,i)=>`${{v.date.slice(5)}}(${{i+1}})`);\n'
        f'    new Chart(document.getElementById(id),{{type:\'bar\',\n'
        f"      data:{{labels,datasets:[{{label:'Likes',data:data.map(v=>v.likes),backgroundColor:'#000',stack:'e'}},{{label:'Comments',data:data.map(v=>v.comments),backgroundColor:'#444',stack:'e'}},{{label:'Shares',data:data.map(v=>v.shares),backgroundColor:'#777',stack:'e'}},{{label:'Saves',data:data.map(v=>v.saves),backgroundColor:'#aaa',stack:'e'}},{{label:'Downloads',data:data.map(v=>v.downloads),backgroundColor:'#ccc',stack:'e'}}]}},\n"
        f'      options:{{responsive:true,interaction:{{mode:\'index\'}},\n'
        f'        plugins:{{legend:{{labels:{{color:\'#333\',boxWidth:10,font:{{size:10,weight:\'500\'}}}}}}}},\n'
        f'        scales:{{x:{{...scaleOpts.x,stacked:true,ticks:{{...scaleOpts.x.ticks,maxRotation:45}}}},y:{{...scaleOpts.y,stacked:true}}}}}}}});\n'
        f'  }}\n'
        f'  {donut_inits}\n'
        f'  {stacked_inits}\n'
        f'}}\n\n'
        f'rTbl();\n'
        f'</script>\n'
        f'</body>\n</html>'
    )

    CAMPAIGNS_DIR.mkdir(exist_ok=True)
    path = CAMPAIGNS_DIR / f"{project['id']}.html"
    path.write_text(html)
    print(f"    Campaign page: {path}")
    return path


# ─── PORTAL INDEX GENERATOR ───────────────────────────────────────────────────
def generate_index(projects_with_data):
    now = datetime.now().strftime("%b %d, %Y %H:%M")

    # Build CAMPAIGNS JS array
    campaign_entries = []
    for p, data in projects_with_data:
        videos   = data.get("videos", []) if data else []
        scraped  = (data.get("scraped_at", "")[:10]) if data else "null"
        tv = sum(v.get("views",0) for v in videos)
        tl = sum(v.get("likes",0) for v in videos)
        ts = sum(v.get("shares",0) for v in videos)
        tc = sum(v.get("comments",0) for v in videos)
        n  = len(videos)
        tags_js = "[" + ",".join(f'"{t}"' for t in p.get("tags",[])) + "]"
        accts_js = "[" + ",".join(f'"{a.lstrip("@")}"' for a in p.get("accounts",[])) + "]"
        scraped_val = f'"{scraped}"' if scraped != "null" else "null"
        entry = (
            "  {\n"
            f'    id: "{p["id"]}",\n'
            f'    name: "{_s(p["name"])}",\n'
            f'    description: "{_s(p.get("description",""))}",\n'
            f'    status: "{p.get("status","active")}",\n'
            f'    accounts: {accts_js},\n'
            f'    tags: {tags_js},\n'
            f'    created: "{p.get("created","")}",\n'
            f'    scraped_at: {scraped_val},\n'
            f'    posts: {n},\n'
            f'    views: {tv},\n'
            f'    likes: {tl},\n'
            f'    comments: {tc},\n'
            f'    shares: {ts},\n'
            "  }"
        )
        campaign_entries.append(entry)

    campaigns_js = "const CAMPAIGNS = [\n" + ",\n".join(campaign_entries) + "\n];"

    # Static CSS for index page
    INDEX_CSS = """  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --black: #000000; --white: #ffffff; --off: #f4f3ef;
    --gray1: #1a1a1a; --gray2: #333333; --gray3: #888888; --gray4: #cccccc; --gray5: #e8e8e4;
    --font: 'Inter', -apple-system, BlinkMacSystemFont, 'Helvetica Neue', Arial, sans-serif;
  }
  html { background: var(--off); }
  body { font-family: var(--font); color: var(--black); min-height: 100vh; }
  a { color: inherit; text-decoration: none; }
  header { background: var(--black); padding: 0 40px; display: flex; align-items: center; justify-content: space-between; height: 64px; border-bottom: 1px solid var(--gray1); }
  .logo { display: flex; align-items: center; gap: 12px; }
  .logo-mark { width: 32px; height: 32px; flex-shrink: 0; }
  .logo-wordmark { font-size: 13px; font-weight: 700; letter-spacing: 0.18em; text-transform: uppercase; color: var(--white); }
  .header-right { font-size: 11px; font-weight: 400; letter-spacing: 0.12em; text-transform: uppercase; color: var(--gray3); }
  .title-bar { background: var(--black); border-bottom: 1px solid #222; padding: 28px 40px 24px; }
  .title-bar-inner { display: flex; align-items: flex-end; justify-content: space-between; gap: 20px; }
  .page-title { font-size: 42px; font-weight: 900; letter-spacing: -0.02em; line-height: 1; color: var(--white); text-transform: uppercase; }
  .page-sub { font-size: 11px; letter-spacing: 0.18em; text-transform: uppercase; color: var(--gray3); margin-top: 8px; font-weight: 400; }
  .new-btn { display: inline-flex; align-items: center; gap: 8px; padding: 12px 22px; border: 1px solid var(--white); color: var(--white); font-family: var(--font); font-size: 11px; font-weight: 600; letter-spacing: 0.15em; text-transform: uppercase; cursor: pointer; background: transparent; transition: background .15s, color .15s; white-space: nowrap; }
  .new-btn:hover { background: var(--white); color: var(--black); }
  .filter-bar { background: var(--black); border-bottom: 1px solid #222; padding: 0 40px; display: flex; align-items: stretch; gap: 0; }
  .filter-tab { padding: 14px 20px; font-size: 11px; font-weight: 500; letter-spacing: 0.14em; text-transform: uppercase; color: var(--gray3); cursor: pointer; border-bottom: 2px solid transparent; transition: color .15s, border-color .15s; white-space: nowrap; }
  .filter-tab:hover { color: var(--white); }
  .filter-tab.active { color: var(--white); border-bottom-color: var(--white); }
  .filter-divider { width: 1px; background: #222; margin: 10px 0; }
  .search-wrap { margin-left: auto; display: flex; align-items: center; }
  .search-box { background: transparent; border: none; border-left: 1px solid #222; padding: 0 20px; height: 100%; font-family: var(--font); font-size: 11px; letter-spacing: 0.1em; text-transform: uppercase; color: var(--white); outline: none; width: 200px; }
  .search-box::placeholder { color: var(--gray3); }
  .summary-strip { background: var(--black); border-bottom: 1px solid #111; padding: 0 40px; display: flex; align-items: stretch; }
  .sum-item { padding: 18px 28px 18px 0; border-right: 1px solid #222; margin-right: 28px; }
  .sum-item:last-child { border-right: none; margin-right: 0; }
  .sum-val { font-size: 24px; font-weight: 900; color: var(--white); letter-spacing: -0.02em; }
  .sum-lbl { font-size: 10px; font-weight: 400; letter-spacing: 0.16em; text-transform: uppercase; color: var(--gray3); margin-top: 2px; }
  .main { padding: 32px 40px 60px; }
  .campaign-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(420px, 1fr)); gap: 1px; background: var(--black); border: 1px solid var(--black); }
  .ccard { background: var(--white); display: flex; flex-direction: column; cursor: pointer; transition: background .15s; position: relative; }
  .ccard:hover { background: var(--off); }
  .ccard-body { padding: 28px 28px 0; flex: 1; }
  .ccard-top { display: flex; align-items: flex-start; gap: 16px; margin-bottom: 16px; }
  .ccard-title { font-size: 17px; font-weight: 800; letter-spacing: -0.01em; line-height: 1.2; flex: 1; }
  .status-chip { font-size: 9px; font-weight: 600; letter-spacing: 0.2em; text-transform: uppercase; padding: 4px 9px; border: 1px solid currentColor; white-space: nowrap; flex-shrink: 0; margin-top: 3px; }
  .status-active   { color: var(--black); border-color: var(--black); }
  .status-pending  { color: var(--gray3); border-color: var(--gray4); }
  .status-complete { color: var(--gray2); border-color: var(--gray3); }
  .status-cancelled{ color: var(--gray4); border-color: var(--gray5); }
  .ccard-accounts { font-size: 11px; font-weight: 400; letter-spacing: 0.08em; color: var(--gray3); margin-bottom: 6px; }
  .ccard-desc { font-size: 12px; color: var(--gray2); line-height: 1.5; margin-bottom: 20px; }
  .ccard-stats { border-top: 1px solid var(--black); display: grid; grid-template-columns: repeat(5, 1fr); }
  .cs { padding: 14px 12px; border-right: 1px solid var(--black); text-align: center; }
  .cs:last-child { border-right: none; }
  .cs-val { font-size: 15px; font-weight: 800; letter-spacing: -0.02em; }
  .cs-lbl { font-size: 9px; font-weight: 400; letter-spacing: 0.14em; text-transform: uppercase; color: var(--gray3); margin-top: 2px; }
  .ccard-foot { border-top: 1px solid var(--black); padding: 10px 28px; display: flex; align-items: center; gap: 8px; background: inherit; }
  .tag { font-size: 9px; font-weight: 500; letter-spacing: 0.14em; text-transform: uppercase; padding: 3px 7px; border: 1px solid var(--gray4); color: var(--gray3); }
  .ccard-date { font-size: 10px; letter-spacing: 0.08em; color: var(--gray4); margin-left: auto; font-weight: 400; text-transform: uppercase; }
  .ccard-arrow { font-size: 14px; color: var(--gray4); transition: color .15s, transform .15s; }
  .ccard:hover .ccard-arrow { color: var(--black); transform: translateX(3px); }
  .ccard-nodata { opacity: 0.45; cursor: default; pointer-events: none; }
  .ccard-pending-badge { position: absolute; top: 0; left: 0; right: 0; height: 2px; background: repeating-linear-gradient(90deg, var(--black) 0, var(--black) 8px, transparent 8px, transparent 14px); }
  .empty { text-align: center; padding: 80px 20px; color: var(--gray3); background: var(--white); }
  .empty-title { font-size: 28px; font-weight: 900; text-transform: uppercase; letter-spacing: 0.05em; color: var(--black); margin-bottom: 8px; }
  .empty-sub { font-size: 12px; letter-spacing: 0.12em; text-transform: uppercase; }
  @media (max-width: 680px) {
    header, .title-bar, .filter-bar, .summary-strip, .main { padding-left: 20px; padding-right: 20px; }
    .page-title { font-size: 28px; }
    .campaign-grid { grid-template-columns: 1fr; }
    .ccard-stats { grid-template-columns: repeat(3, 1fr); }
  }"""

    # Static JS rendering functions for index
    INDEX_JS_STATIC = """function fmt(n) {
  if (n >= 1000000) return (n/1000000).toFixed(1) + 'M';
  if (n >= 1000)    return (n/1000).toFixed(1) + 'K';
  return n.toString();
}
function statusClass(s) {
  return {active:'status-active',pending:'status-pending',complete:'status-complete',cancelled:'status-cancelled'}[s]||'status-pending';
}
function renderGrid(data) {
  const grid = document.getElementById('campaignGrid');
  if (!data.length) {
    grid.innerHTML = '<div class="empty"><div class="empty-title">No Campaigns</div><div class="empty-sub">Add a .json config to /projects/ and run scraper.py</div></div>';
    return;
  }
  grid.innerHTML = data.map(c => {
    const hasData = c.posts > 0;
    const link    = hasData ? `campaigns/${c.id}.html` : '#';
    const sc      = statusClass(c.status);
    const accts   = c.accounts.slice(0,3).map(a=>'@'+a).join(' · ') + (c.accounts.length>3?' + more':'');
    const tagHTML = c.tags.slice(0,3).map(t=>`<span class="tag">${t}</span>`).join('');
    return `<a class="ccard${hasData?'':' ccard-nodata'}" href="${link}">
      ${!hasData?'<div class="ccard-pending-badge"></div>':''}
      <div class="ccard-body">
        <div class="ccard-top">
          <div class="ccard-title">${c.name}</div>
          <span class="status-chip ${sc}">${c.status}</span>
        </div>
        <div class="ccard-accounts">${accts}</div>
        ${c.description?`<div class="ccard-desc">${c.description}</div>`:''}
      </div>
      <div class="ccard-stats">
        <div class="cs"><div class="cs-val">${c.posts}</div><div class="cs-lbl">Posts</div></div>
        <div class="cs"><div class="cs-val">${fmt(c.views)}</div><div class="cs-lbl">Views</div></div>
        <div class="cs"><div class="cs-val">${fmt(c.likes)}</div><div class="cs-lbl">Likes</div></div>
        <div class="cs"><div class="cs-val">${fmt(c.shares)}</div><div class="cs-lbl">Shares</div></div>
        <div class="cs"><div class="cs-val">${c.accounts.length}</div><div class="cs-lbl">Accounts</div></div>
      </div>
      <div class="ccard-foot">
        ${tagHTML}
        <span class="ccard-date">${c.scraped_at?'Updated '+c.scraped_at:'Not yet scraped'}</span>
        ${hasData?'<span class="ccard-arrow">→</span>':''}
      </div>
    </a>`;
  }).join('');
}
function updateSummary(data) {
  const active = data.filter(c=>c.status==='active').length;
  const posts  = data.reduce((s,c)=>s+c.posts,0);
  const views  = data.reduce((s,c)=>s+c.views,0);
  document.getElementById('s-total').textContent  = data.length;
  document.getElementById('s-active').textContent = active;
  document.getElementById('s-posts').textContent  = posts.toLocaleString();
  document.getElementById('s-views').textContent  = fmt(views);
  document.getElementById('campaignCount').textContent = `${data.length} campaign${data.length!==1?'s':''} · ${active} active`;
}
let currentFilter='all', currentSearch='';
function filterCards(status,el) {
  currentFilter=status;
  document.querySelectorAll('.filter-tab').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');
  applyFilters();
}
function searchCards(q) {
  currentSearch=q.toLowerCase();
  applyFilters();
}
function applyFilters() {
  let data=CAMPAIGNS;
  if(currentFilter!=='all') data=data.filter(c=>c.status===currentFilter);
  if(currentSearch) data=data.filter(c=>
    c.name.toLowerCase().includes(currentSearch)||
    c.accounts.some(a=>a.toLowerCase().includes(currentSearch))||
    (c.description||'').toLowerCase().includes(currentSearch)
  );
  renderGrid(data);
}
updateSummary(CAMPAIGNS);
renderGrid(CAMPAIGNS);"""

    html = (
        '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        '<title>Interlude Studios — Fan Page Database</title>\n'
        f'{FONT_TAG}\n'
        f'<style>\n{INDEX_CSS}\n</style>\n'
        '</head>\n<body>\n\n'
        # Header
        '<header>\n'
        '  <div class="logo">\n'
        f'    <div class="logo-mark">{LOGO_SVG}</div>\n'
        '    <span class="logo-wordmark">Interlude Studios</span>\n'
        '  </div>\n'
        f'  <div class="header-right">Fan Page Database · {now}</div>\n'
        '</header>\n\n'
        # Title bar
        '<div class="title-bar">\n'
        '  <div class="title-bar-inner">\n'
        '    <div>\n'
        '      <div class="page-title">Campaigns +</div>\n'
        '      <div class="page-sub" id="campaignCount">Loading…</div>\n'
        '    </div>\n'
        '    <button class="new-btn" onclick="alert(\'To add a campaign: create a new .json file in the /projects/ folder and re-run scraper.py\')">\n'
        '      + New Campaign\n'
        '    </button>\n'
        '  </div>\n'
        '</div>\n\n'
        # Filter bar
        '<div class="filter-bar">\n'
        '  <div class="filter-tab active" onclick="filterCards(\'all\',this)">All</div>\n'
        '  <div class="filter-tab" onclick="filterCards(\'active\',this)">Active</div>\n'
        '  <div class="filter-tab" onclick="filterCards(\'pending\',this)">Pending</div>\n'
        '  <div class="filter-tab" onclick="filterCards(\'complete\',this)">Complete</div>\n'
        '  <div class="filter-tab" onclick="filterCards(\'cancelled\',this)">Cancelled</div>\n'
        '  <div class="search-wrap">\n'
        '    <input class="search-box" placeholder="Search campaigns…" oninput="searchCards(this.value)">\n'
        '  </div>\n'
        '</div>\n\n'
        # Summary strip
        '<div class="summary-strip">\n'
        '  <div class="sum-item"><div class="sum-val" id="s-total">—</div><div class="sum-lbl">Total Campaigns</div></div>\n'
        '  <div class="sum-item"><div class="sum-val" id="s-active">—</div><div class="sum-lbl">Active</div></div>\n'
        '  <div class="sum-item"><div class="sum-val" id="s-posts">—</div><div class="sum-lbl">Total Posts</div></div>\n'
        '  <div class="sum-item"><div class="sum-val" id="s-views">—</div><div class="sum-lbl">Total Views</div></div>\n'
        '</div>\n\n'
        # Main
        '<div class="main">\n'
        '  <div class="campaign-grid" id="campaignGrid"></div>\n'
        '</div>\n\n'
        # Script
        '<script>\n'
        f'{campaigns_js}\n\n'
        f'{INDEX_JS_STATIC}\n'
        '</script>\n'
        '</body>\n</html>'
    )

    (ROOT / "index.html").write_text(html)
    print(f"  Portal index generated: {ROOT / 'index.html'}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Interlude Studios Fan Page Portal — Scraper & Generator")
    parser.add_argument("--project",       help="Scrape only this project ID")
    parser.add_argument("--generate-only", action="store_true", help="Rebuild HTML without scraping")
    parser.add_argument("--export-xlsx",   action="store_true", help="Export Excel files")
    args = parser.parse_args()

    print("★ Interlude Studios Fan Page Portal")
    print(f"   Root: {ROOT}")
    print()

    all_projects = load_all_projects()
    if not all_projects:
        print("No projects found in /projects/. Add a JSON config and re-run.")
        return

    to_process = [p for p in all_projects if not args.project or p["id"] == args.project]
    print(f"  Found {len(all_projects)} project(s), processing {len(to_process)}")
    print()

    projects_with_data = []

    for p in to_process:
        print(f"▶ {p['name']} [{p.get('status','active')}]")

        if args.generate_only:
            data = load_project_data(p["id"])
            if not data:
                print(f"  No cached data for {p['id']}, skipping HTML generation.")
                projects_with_data.append((p, None))
                continue
        else:
            if p.get("status") == "cancelled":
                print(f"  Skipping (cancelled)")
                data = load_project_data(p["id"])
                projects_with_data.append((p, data))
                continue
            print(f"  Scraping {len(p['accounts'])} account(s)...")
            videos = scrape_project(p)
            if not videos:
                print(f"  No videos found — check accounts and API key.")
                projects_with_data.append((p, None))
                continue
            print(f"  Saving data ({len(videos)} total videos)...")
            save_project_data(p["id"], videos)
            data = {"scraped_at": datetime.now().isoformat(), "videos": videos}
            if args.export_xlsx:
                print(f"  Exporting Excel...")
                export_xlsx(p, videos)

        print(f"  Generating campaign dashboard...")
        generate_campaign_html(p, data)
        projects_with_data.append((p, data))
        print()

    # Include unprocessed projects in the index
    processed_ids = {p["id"] for p, _ in projects_with_data}
    for p in all_projects:
        if p["id"] not in processed_ids:
            data = load_project_data(p["id"])
            projects_with_data.append((p, data))

    print("  Regenerating portal index...")
    generate_index(projects_with_data)
    print()
    print("✓ Done!")

if __name__ == "__main__":
    main()
